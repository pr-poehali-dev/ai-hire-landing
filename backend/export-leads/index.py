import json
import os
from typing import Dict, Any
import psycopg2
from psycopg2.extras import RealDictCursor
from io import BytesIO
import base64

def handler(event: Dict[str, Any], context: Any) -> Dict[str, Any]:
    '''
    Экспорт лидов в Excel формат
    GET /export?format=excel - экспортировать все лиды
    GET /export?format=excel&stage_id=1 - экспортировать лиды конкретного этапа
    POST /export - экспортировать с фильтрами (priority, source, date_from, date_to)
    '''
    method: str = event.get('httpMethod', 'GET')
    
    if method == 'OPTIONS':
        return {
            'statusCode': 200,
            'headers': {
                'Access-Control-Allow-Origin': '*',
                'Access-Control-Allow-Methods': 'GET, POST, OPTIONS',
                'Access-Control-Allow-Headers': 'Content-Type, X-Auth-Token',
                'Access-Control-Max-Age': '86400'
            },
            'body': '',
            'isBase64Encoded': False
        }
    
    dsn = os.environ['DATABASE_URL']
    conn = psycopg2.connect(dsn)
    cursor = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        params = event.get('queryStringParameters') or {}
        
        if method == 'POST':
            body_data = json.loads(event.get('body', '{}'))
            params.update(body_data)
        
        query = '''
            SELECT 
                l.id,
                l.name,
                l.phone,
                l.email,
                l.company,
                l.vacancy,
                l.source,
                l.priority,
                l.notes,
                l.created_at,
                l.updated_at,
                s.name as stage_name,
                (SELECT COUNT(*) FROM lead_tasks WHERE lead_id = l.id AND completed = false) as open_tasks,
                (SELECT COUNT(*) FROM lead_tasks WHERE lead_id = l.id AND completed = true) as completed_tasks,
                (SELECT COUNT(*) FROM lead_comments WHERE lead_id = l.id) as comments_count,
                (SELECT COUNT(*) FROM lead_calls WHERE lead_id = l.id) as calls_count
            FROM lead_data l
            LEFT JOIN lead_stages s ON l.stage_id = s.id
            WHERE 1=1
        '''
        
        query_params = []
        
        if params.get('stage_id'):
            query += ' AND l.stage_id = %s'
            query_params.append(params['stage_id'])
        
        if params.get('priority'):
            query += ' AND l.priority = %s'
            query_params.append(params['priority'])
        
        if params.get('source'):
            query += ' AND l.source = %s'
            query_params.append(params['source'])
        
        if params.get('date_from'):
            query += ' AND l.created_at >= %s'
            query_params.append(params['date_from'])
        
        if params.get('date_to'):
            query += ' AND l.created_at <= %s'
            query_params.append(params['date_to'])
        
        query += ' ORDER BY l.created_at DESC'
        
        cursor.execute(query, query_params)
        leads = cursor.fetchall()
        
        if params.get('format') == 'excel':
            excel_data = generate_excel(leads)
            
            return {
                'statusCode': 200,
                'headers': {
                    'Access-Control-Allow-Origin': '*',
                    'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    'Content-Disposition': 'attachment; filename="leads_export.xlsx"'
                },
                'body': excel_data,
                'isBase64Encoded': True
            }
        else:
            return {
                'statusCode': 200,
                'headers': {
                    'Access-Control-Allow-Origin': '*',
                    'Content-Type': 'application/json'
                },
                'body': json.dumps({
                    'success': True,
                    'leads': [dict(l) for l in leads],
                    'total': len(leads)
                }, default=str),
                'isBase64Encoded': False
            }
    
    finally:
        cursor.close()
        conn.close()


def generate_excel(leads: list) -> str:
    '''Генерация Excel файла из данных лидов'''
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Лиды"
        
        headers = [
            'ID', 'Имя', 'Телефон', 'Email', 'Компания', 'Вакансия', 
            'Источник', 'Приоритет', 'Этап', 'Открытых задач', 
            'Завершенных задач', 'Комментариев', 'Звонков', 
            'Создан', 'Обновлен', 'Заметки'
        ]
        
        header_fill = PatternFill(start_color='3B82F6', end_color='3B82F6', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for row_idx, lead in enumerate(leads, 2):
            ws.cell(row=row_idx, column=1, value=lead['id'])
            ws.cell(row=row_idx, column=2, value=lead['name'])
            ws.cell(row=row_idx, column=3, value=lead['phone'])
            ws.cell(row=row_idx, column=4, value=lead.get('email', ''))
            ws.cell(row=row_idx, column=5, value=lead.get('company', ''))
            ws.cell(row=row_idx, column=6, value=lead.get('vacancy', ''))
            ws.cell(row=row_idx, column=7, value=lead['source'])
            ws.cell(row=row_idx, column=8, value=lead['priority'])
            ws.cell(row=row_idx, column=9, value=lead.get('stage_name', ''))
            ws.cell(row=row_idx, column=10, value=lead.get('open_tasks', 0))
            ws.cell(row=row_idx, column=11, value=lead.get('completed_tasks', 0))
            ws.cell(row=row_idx, column=12, value=lead.get('comments_count', 0))
            ws.cell(row=row_idx, column=13, value=lead.get('calls_count', 0))
            ws.cell(row=row_idx, column=14, value=str(lead['created_at']))
            ws.cell(row=row_idx, column=15, value=str(lead.get('updated_at', '')))
            ws.cell(row=row_idx, column=16, value=lead.get('notes', ''))
        
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 15
        
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['P'].width = 40
        
        output = BytesIO()
        wb.save(output)
        excel_bytes = output.getvalue()
        
        return base64.b64encode(excel_bytes).decode('utf-8')
        
    except ImportError:
        return base64.b64encode(b'Error: openpyxl not installed').decode('utf-8')
